import os
import io
import re
import json  # Добавляем json
from pathlib import Path  # Для работы с путями
from dotenv import load_dotenv
from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Literal
import pandas as pd
import google.generativeai as genai
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from utils import extract_text_from_pdf, parse_gemini_response, DEFAULT_PROMPT_TEMPLATE

# --- Локализация ---
BASE_DIR = Path(__file__).resolve().parent
LOCALES_DIR = BASE_DIR / "locales"
SUPPORTED_LANGUAGES = Literal["ru", "en"]
DEFAULT_LANGUAGE: SUPPORTED_LANGUAGES = "ru"

translations: Dict[str, Dict[str, str]] = {}


def load_translations():
    for lang_file in LOCALES_DIR.glob("*.json"):
        lang_code = lang_file.stem  # ru, en
        if (
            lang_code in SUPPORTED_LANGUAGES.__args__
        ):  # Проверка, что язык поддерживается
            with open(lang_file, "r", encoding="utf-8") as f:
                translations[lang_code] = json.load(f)
    if DEFAULT_LANGUAGE not in translations:
        raise RuntimeError(
            f"Default language '{DEFAULT_LANGUAGE}' translations not found in {LOCALES_DIR}"
        )
    print(f"Loaded translations for: {list(translations.keys())}")


load_translations()  # Загружаем переводы при старте приложения


def get_locale_strings(lang: Optional[str] = None) -> Dict[str, str]:
    if lang and lang in translations:
        return translations[lang]
    return translations[DEFAULT_LANGUAGE]


# --- Конец блока локализации ---


load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

if not GEMINI_API_KEY:
    raise ValueError("GEMINI_API_KEY не найден в .env файле.")

genai.configure(api_key=GEMINI_API_KEY)

GENERATION_CONFIG = {
    "temperature": 0.7,
    "top_p": 1,
    "top_k": 1,
    "max_output_tokens": 8192,
}
SAFETY_SETTINGS = [
    {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
    {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
    {
        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
        "threshold": "BLOCK_MEDIUM_AND_ABOVE",
    },
    {
        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
        "threshold": "BLOCK_MEDIUM_AND_ABOVE",
    },
]

try:
    model = genai.GenerativeModel(
        model_name="gemini-1.5-flash-latest",
        generation_config=GENERATION_CONFIG,
        safety_settings=SAFETY_SETTINGS,
    )
    print("Using model: gemini-1.5-flash-latest")
except Exception as e:
    print(f"Failed to load gemini-1.5-flash-latest, trying gemini-pro. Error: {e}")
    try:
        model = genai.GenerativeModel(
            model_name="gemini-pro",
            generation_config=GENERATION_CONFIG,
            safety_settings=SAFETY_SETTINGS,
        )
        print("Using model: gemini-pro")
    except Exception as e_pro:
        raise RuntimeError(f"Could not initialize any Gemini model. Error: {e_pro}")

app = FastAPI(title="Генератор Тест-кейсов на AI")
app.mount(
    "/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static"
)  # Используем BASE_DIR
templates = Jinja2Templates(
    directory=str(BASE_DIR / "templates")
)  # Используем BASE_DIR


class TestCase(BaseModel):
    Название: str
    Шаги: str
    Ожидаемый_результат: str = Field(..., alias="Ожидаемый результат")
    Тип: str


class ExportRequest(BaseModel):
    test_cases: List[TestCase]


@app.get("/", response_class=HTMLResponse)
async def get_index(
    request: Request, lang: Optional[SUPPORTED_LANGUAGES] = Query(DEFAULT_LANGUAGE)
):
    # lang будет автоматически валидироваться благодаря Literal
    locale_strings = get_locale_strings(lang)
    # Передаем текущий язык и все строки локализации в шаблон
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "lang": lang,
            "locales": locale_strings,
            "supported_languages": SUPPORTED_LANGUAGES.__args__,
        },
    )


@app.post("/generate")
async def generate_test_cases_endpoint(
    request: Request,
    requirements_text: Optional[str] = Form(None),
    pdf_file: Optional[UploadFile] = File(None),
    custom_prompt: Optional[str] = Form(None),
    lang: Optional[SUPPORTED_LANGUAGES] = Form(
        DEFAULT_LANGUAGE
    ),  # Получаем язык из формы
):
    loc = get_locale_strings(lang)  # Получаем строки для текущего языка
    input_text = ""
    if pdf_file:
        if pdf_file.content_type != "application/pdf":
            raise HTTPException(
                status_code=400,
                detail=loc.get("backend_error_invalid_file_type", "Invalid file type."),
            )
        try:
            contents = await pdf_file.read()
            pdf_stream = io.BytesIO(contents)
            extracted_text = extract_text_from_pdf(pdf_stream)
            if not extracted_text or not extracted_text.strip():
                raise HTTPException(
                    status_code=400,
                    detail=loc.get(
                        "backend_error_pdf_extraction_failed", "PDF extraction failed."
                    ),
                )
            input_text = extracted_text
        except Exception as e:
            # Не используем loc здесь, так как это более общая ошибка сервера
            raise HTTPException(
                status_code=500, detail=f"Ошибка при обработке PDF: {e}"
            )
        finally:
            await pdf_file.close()
    elif requirements_text:
        input_text = requirements_text.strip()
    else:
        raise HTTPException(
            status_code=400,
            detail=loc.get("backend_error_no_input_provided", "No input provided."),
        )

    if not input_text:
        raise HTTPException(
            status_code=400,
            detail=loc.get("backend_error_input_empty", "Input is empty."),
        )

    MAX_INPUT_CHARS = 60000
    if len(input_text) > MAX_INPUT_CHARS:
        input_text = input_text[:MAX_INPUT_CHARS]
        print(f"Warning: Input text was truncated to {MAX_INPUT_CHARS} characters.")

    prompt_template_to_use = (
        custom_prompt
        if custom_prompt and "{requirements_text}" in custom_prompt
        else DEFAULT_PROMPT_TEMPLATE
    )
    # Важно: DEFAULT_PROMPT_TEMPLATE тоже нужно бы локализовать или сделать так,
    # чтобы он был нейтральным к языку и Gemini генерировал на языке требований.
    # Пока оставим его как есть, т.к. он описывает структуру.
    final_prompt = prompt_template_to_use.format(requirements_text=input_text)

    try:
        print(f"Sending prompt to Gemini (first 100 chars): {final_prompt[:100]}")
        response = model.generate_content(final_prompt)
        generated_text = ""
        if hasattr(response, "text") and response.text:
            generated_text = response.text
        elif hasattr(response, "parts") and response.parts:
            generated_text = "".join(
                part.text for part in response.parts if hasattr(part, "text")
            )
        elif isinstance(response, str):
            generated_text = response

        if not generated_text.strip():
            if hasattr(response, "prompt_feedback") and response.prompt_feedback:
                for rating in response.prompt_feedback.safety_ratings:
                    if rating.blocked:
                        error_message = loc.get(
                            "backend_error_gemini_request_blocked",
                            "Request blocked due to safety settings: {category}. Try modifying the text.",
                        )
                        raise HTTPException(
                            status_code=400,
                            detail=error_message.format(category=rating.category),
                        )
            raise HTTPException(
                status_code=500,
                detail=loc.get(
                    "backend_error_gemini_text_response_failed",
                    "Failed to get text response from Gemini.",
                ),
            )

    except Exception as e:
        print(f"Error calling Gemini API or processing response: {e}")
        if "API key not valid" in str(e):
            raise HTTPException(
                status_code=500,
                detail=loc.get(
                    "backend_error_gemini_key_invalid", "Gemini API key error."
                ),
            )
        elif "Deadline" in str(e) or "timeout" in str(e):
            raise HTTPException(
                status_code=504,
                detail=loc.get(
                    "backend_error_gemini_timeout", "Timeout contacting Gemini API."
                ),
            )
        if hasattr(e, "message"):
            error_detail_msg = e.message
        else:
            error_detail_msg = str(e)
        error_message = loc.get(
            "backend_error_gemini_api_error", "Error calling Gemini API: {error_detail}"
        )
        raise HTTPException(
            status_code=500, detail=error_message.format(error_detail=error_detail_msg)
        )

    test_cases = parse_gemini_response(
        generated_text
    )  # parse_gemini_response должен быть нечувствителен к языку структуры

    if not test_cases:
        return JSONResponse(
            content={
                "test_cases": [],
                "message": loc.get(
                    "backend_error_parsing_failed", "Could not parse test cases."
                ),
            },
            status_code=200,
        )
    return {
        "test_cases": test_cases
    }  # Тест-кейсы будут на том языке, на котором их сгенерировал Gemini


# Функции экспорта остаются без изменений в логике локализации,
# так как они экспортируют уже полученные данные. Заголовки файлов будут на английском.


@app.post("/export_csv")
async def export_to_csv(payload: ExportRequest):
    # ... (код как был, только исправления регулярных выражений)
    test_cases_data = [tc.model_dump(by_alias=True) for tc in payload.test_cases]
    if not test_cases_data:
        # loc = get_locale_strings(DEFAULT_LANGUAGE) # Или получить язык из запроса, если это важно для этой ошибки
        raise HTTPException(
            status_code=400, detail="No data for export."
        )  # Простое сообщение

    for tc_item in test_cases_data:
        if "Шаги" in tc_item and isinstance(tc_item["Шаги"], str):
            steps_str = tc_item["Шаги"]
            steps_str = re.sub(r"(?<=[^\n])\s+(?=\d+\.\s)", r"\n", steps_str)
            steps_str = re.sub(
                r"^(?!\s*\d+\.\s)([^\n]*?)\s*(1\.\s)",
                r"\1\n\2",
                steps_str,
                count=1,
                flags=re.MULTILINE,
            )
            tc_item["Шаги"] = steps_str.replace("\n", "; ")

    df = pd.DataFrame(test_cases_data)
    stream = io.StringIO()
    df.to_csv(stream, index=False, encoding="utf-8-sig")

    response = StreamingResponse(
        iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=test_cases.csv"},
    )
    return response


@app.post("/export_excel")
async def export_to_excel(payload: ExportRequest):
    # ... (код как был, только исправления регулярных выражений)
    test_cases_data = [tc.model_dump(by_alias=True) for tc in payload.test_cases]
    if not test_cases_data:
        # loc = get_locale_strings(DEFAULT_LANGUAGE)
        raise HTTPException(
            status_code=400, detail="No data for export."
        )  # Простое сообщение

    for item in test_cases_data:
        if "Шаги" in item and isinstance(item["Шаги"], str):
            steps_str = item["Шаги"]
            item["Шаги"] = re.sub(r"(?<=[^\n])\s+(?=\d+\.\s)", r"\n", steps_str)
            item["Шаги"] = re.sub(
                r"^(?!\s*\d+\.\s)([^\n]*?)\s*(1\.\s)",
                r"\1\n\2",
                item["Шаги"],
                count=1,
                flags=re.MULTILINE,
            )

    df = pd.DataFrame(test_cases_data)
    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        df.to_excel(
            writer, index=False, sheet_name="Test-Cases"
        )  # Sheet name in English
        workbook = writer.book
        worksheet = writer.sheets["Test-Cases"]

        steps_col_letter = None
        steps_col_idx_df_coord = -1
        try:
            steps_col_idx_df_coord = df.columns.get_loc("Шаги")
            steps_col_letter = get_column_letter(steps_col_idx_df_coord + 1)
        except KeyError:
            print("Warning: Column 'Шаги' not found in DataFrame for Excel styling.")

        if steps_col_letter and steps_col_idx_df_coord != -1:
            # Индексы openpyxl для колонки начинаются с 1
            openpyxl_col_idx = steps_col_idx_df_coord + 1
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=openpyxl_col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            max_width = 0
            for row_cells in worksheet.iter_rows(
                min_col=openpyxl_col_idx,
                max_col=openpyxl_col_idx,
                min_row=1,
                max_row=worksheet.max_row,
            ):
                for cell in row_cells:
                    if cell.value:
                        cell_content = str(cell.value)
                        lines = cell_content.split("\n")
                        for line in lines:
                            if len(line) > max_width:
                                max_width = len(line)
            adjusted_width = min(max_width + 5, 70)
            if adjusted_width > 10:
                worksheet.column_dimensions[steps_col_letter].width = adjusted_width
    stream.seek(0)
    response = StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_cases.xlsx"},
    )
    return response


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
